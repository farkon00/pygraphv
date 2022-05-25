import openpyxl

wb = openpyxl.load_workbook('Attributes.xlsx')
sheet = wb.worksheets[int(input("Sheet index: "))]

"""for j, i in enumerate(sheet.rows):
    if i[2].value not in ("int", "str", "float", "bool"):
        print(i[2].value)
        i[2].value = 'str'"""

"""
res = openpyxl.Workbook()
main_sheet = res.active
g_sheet = res.create_sheet("G")
n_sheet = res.create_sheet("N")
c_sheet = res.create_sheet("C")
e_sheet = res.create_sheet("E")

for i in sheet.rows:
    if len(i[1].value) >= 4: main_sheet.append([i[0].value, i[1].value, i[2].value]); continue
    if "G" in i[1].value: g_sheet.append([i[0].value, i[1].value, i[2].value])
    if "N" in i[1].value: n_sheet.append([i[0].value, i[1].value, i[2].value])
    if "C" in i[1].value: c_sheet.append([i[0].value, i[1].value, i[2].value])
    if "E" in i[1].value: e_sheet.append([i[0].value, i[1].value, i[2].value])"""

f = open("ResStyles.py", "w")
array = []
for i in sheet.rows:
    array.append(i[0].value)
    f.write(f"{i[0].value}: {i[2].value} | None = None\n")

print(array)
f.close()